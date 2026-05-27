"""
app.py

Main Flask web application file.

Responsibilities:
1. Display login page
2. Authenticate user from role_login table
3. Create login session
4. Redirect user to role-based dashboard
5. All four role dashboards are wired and active
6. Admin panel API routes (superadmin + admin only):
   - Add user to role_login            (superadmin only)
   - Add / upsert client in clients    (superadmin + admin)
   - Bulk upsert clients from .xlsx    (superadmin + admin)
   - Add server to server_info         (superadmin + admin)
   - Bulk insert servers from .xlsx    (superadmin + admin, skips duplicates)
   - Download .xlsx templates

Login source table : role_login
Login fields       : email, password
Supported roles    : superadmin, admin, data, operator
Admin Panel access : superadmin, admin
"""

from __future__ import annotations

import io
import os
from functools import wraps
from typing import Callable, Optional

import openpyxl
import pymysql
from dotenv import load_dotenv
from flask import (
    Flask,
    jsonify,
    render_template,
    request,
    redirect,
    send_file,
    url_for,
    session,
)

from helper import log_info, log_error, log_warning, log_update, log_added


MODULE_NAME = "app"


# =========================
# Flask App Setup
# =========================

load_dotenv()

app = Flask(__name__)

app.secret_key = os.getenv(
    "SECRET_KEY",
    "change-this-secret-key-in-env-file",
)


# =========================
# Constants
# =========================

VALID_ROLES: frozenset[str] = frozenset({"superadmin", "admin", "data", "operator"})
ADMIN_PANEL_ROLES: frozenset[str] = frozenset({"superadmin", "admin"})

# ── Client ────────────────────────────────────────────────────────────────────

# Client columns in form and bulk upload (excludes 'server')
CLIENT_FORM_COLUMNS: list[str] = [
    "userId",
    "alias",
    "Broker",
    "algo",
    "Running Type",
    "Category",
    "SubCategory",
    "Acc Type",
]

# Columns that MUST have a value (single-add and bulk)
CLIENT_REQUIRED_FIELDS: frozenset[str] = frozenset({"userId", "alias", "Broker"})

# Bulk upload: only these columns must be present as headers in the file
CLIENT_REQUIRED_BULK_COLUMNS: list[str] = ["userId", "alias", "Broker"]

CLIENT_MAX_LENGTHS: dict[str, int] = {
    "userId": 20,
    "alias": 30,
    "Broker": 20,
    "algo": 10,
    "Running Type": 20,
    "Category": 20,
    "SubCategory": 20,
    "Acc Type": 20,
}

# ── Server ────────────────────────────────────────────────────────────────────

# Server columns in form and bulk upload.
# Excludes: Dte, Aum, Remarks, Operator, Stoxxo URL
SERVER_FORM_COLUMNS: list[str] = [
    "Server",
    "Username",
    "IP",
    "Password",
    "Stoxxo Id",
    "Stoxxo Password",
    "Algo",
    "Expiry",
    "Subscriptions",
    "Logins",
    "Active",
    "Avlbl",
]

# Columns that MUST have a value (single-add and bulk)
SERVER_REQUIRED_FIELDS: frozenset[str] = frozenset({
    "Server", "Username", "IP", "Password", "Stoxxo Id", "Stoxxo Password"
})

# Bulk upload: only these columns must be present as headers in the file
SERVER_REQUIRED_BULK_COLUMNS: list[str] = [
    "Server", "Username", "IP", "Password", "Stoxxo Id", "Stoxxo Password"
]

# INT-typed server columns — optional, default 0 when absent/empty
SERVER_INT_COLUMNS: frozenset[str] = frozenset({"Subscriptions", "Logins", "Active", "Avlbl"})

# VARCHAR server columns with max lengths
SERVER_VARCHAR_MAX_LENGTHS: dict[str, int] = {
    "Server": 10,
    "Username": 20,
    "IP": 20,
    "Password": 30,
    "Stoxxo Id": 30,
    "Stoxxo Password": 30,
    "Algo": 10,
    "Expiry": 20,
}

# ── Role Login ────────────────────────────────────────────────────────────────

ROLE_LOGIN_MAX_LENGTHS: dict[str, int] = {
    "role": 20,
    "name": 30,
    "ops_name": 20,
    "email": 30,
    "password": 30,
}


# =========================
# MySQL Config
# =========================

def get_mysql_config() -> dict:
    """Reads MySQL configuration from environment variables."""

    mysql_port_raw = os.getenv("MYSQL_PORT", "3306")

    try:
        mysql_port = int(mysql_port_raw)
    except ValueError as exc:
        raise ValueError("MYSQL_PORT must be a valid integer") from exc

    return {
        "host": os.getenv("MYSQL_HOST", "localhost"),
        "port": mysql_port,
        "user": os.getenv("MYSQL_USER", "root"),
        "password": os.getenv("MYSQL_PASSWORD", ""),
        "database": os.getenv("MYSQL_DATABASE", "cmp"),
    }


def get_db_connection() -> pymysql.connections.Connection:
    """
    Creates and returns a MySQL database connection.
    Caller is responsible for closing it.
    """

    config = get_mysql_config()

    return pymysql.connect(
        host=config["host"],
        port=config["port"],
        user=config["user"],
        password=config["password"],
        database=config["database"],
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
    )


# =========================
# Auth Helpers
# =========================

def get_user_by_email(email: str) -> Optional[dict]:
    """Fetches a user row from role_login by email."""

    query = """
        SELECT role, name, ops_name, email, password
        FROM role_login
        WHERE email = %s
        LIMIT 1
    """

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query, (email,))
            return cursor.fetchone()
    finally:
        connection.close()


def is_valid_role(role: str) -> bool:
    return role in VALID_ROLES


def login_required(view_func: Callable) -> Callable:
    """Decorator: redirects unauthenticated users to login."""

    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("is_logged_in"):
            log_warning(
                module=MODULE_NAME,
                action="login_required",
                message="Unauthorized access attempt. Redirecting to login.",
                status="FAILED",
            )
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapper


def role_required(required_role: str) -> Callable:
    """Decorator: returns HTTP 403 HTML page if role does not match."""

    def decorator(view_func: Callable) -> Callable:
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            current_role = session.get("role")

            if current_role != required_role:
                log_warning(
                    module=MODULE_NAME,
                    action="role_required",
                    message=(
                        f"Access denied. Required: {required_role}, "
                        f"current: {current_role}"
                    ),
                    status="FAILED",
                )
                return render_template(
                    "login.html",
                    error="You do not have permission to access this page.",
                ), 403

            return view_func(*args, **kwargs)

        return wrapper

    return decorator


def redirect_user_by_role(role: str):
    """Redirects authenticated user to their role-specific dashboard."""

    if role == "superadmin":
        return redirect(url_for("superadmin_dashboard"))
    if role == "admin":
        return redirect(url_for("admin_dashboard"))
    if role == "data":
        return redirect(url_for("data_dashboard"))
    if role == "operator":
        return redirect(url_for("operator_dashboard"))

    return render_template("login.html", error="Invalid user role.")


# =========================
# Admin Panel Helpers
# =========================

def check_admin_panel_access() -> Optional[tuple]:
    """
    Checks if the current session role has admin panel access.

    Returns:
        None if access is granted.
        JSON error tuple (Response, status_code) if denied.
    """

    if session.get("role") not in ADMIN_PANEL_ROLES:
        return jsonify({"success": False, "error": "Permission denied."}), 403

    return None


def get_allowed_new_roles(current_role: str) -> list[str]:
    """
    Returns roles this user may assign when adding a new user.
    superadmin: all four roles.
    admin: data and operator only.
    """

    if current_role == "superadmin":
        return ["superadmin", "admin", "data", "operator"]
    if current_role == "admin":
        return ["data", "operator"]
    return []


def validate_string_field(value: str, field_name: str, max_length: int) -> Optional[str]:
    """
    Validates a required string field against max length.

    Returns None if valid, or an error message string.
    """

    stripped = str(value).strip()

    if not stripped:
        return f"Field '{field_name}' is required."

    if len(stripped) > max_length:
        return (
            f"Field '{field_name}' exceeds maximum length "
            f"of {max_length} characters."
        )

    return None


# =========================
# DB: User Operations
# =========================

def user_exists_by_ops_name_or_email(
    connection: pymysql.connections.Connection,
    ops_name: str,
    email: str,
) -> Optional[dict]:
    """Returns a matching row if ops_name or email already exists, else None."""

    query = """
        SELECT ops_name, email
        FROM role_login
        WHERE ops_name = %s OR email = %s
        LIMIT 1
    """

    with connection.cursor() as cursor:
        cursor.execute(query, (ops_name, email))
        return cursor.fetchone()


def insert_role_login_user(
    connection: pymysql.connections.Connection,
    user_data: dict,
) -> None:
    """Inserts a new user into role_login."""

    query = """
        INSERT INTO role_login (role, name, ops_name, email, password)
        VALUES (%s, %s, %s, %s, %s)
    """

    values = (
        user_data["role"].strip(),
        user_data["name"].strip(),
        user_data["ops_name"].strip(),
        user_data["email"].strip(),
        user_data["password"].strip(),
    )

    with connection.cursor() as cursor:
        cursor.execute(query, values)


# =========================
# DB: Client Operations
# =========================

def client_exists_by_user_id(
    connection: pymysql.connections.Connection,
    user_id: str,
) -> bool:
    """Returns True if a client with the given userId already exists."""

    query = "SELECT userId FROM clients WHERE userId = %s LIMIT 1"

    with connection.cursor() as cursor:
        cursor.execute(query, (user_id,))
        return cursor.fetchone() is not None


def insert_client(
    connection: pymysql.connections.Connection,
    client_data: dict,
) -> None:
    """Inserts a new client row."""

    query = """
        INSERT INTO clients
            (userId, alias, Broker, algo, `Running Type`, `Operator Name`,
             Category, SubCategory, `Acc Type`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    values = (
        str(client_data.get("userId", "")).strip(),
        str(client_data.get("alias", "")).strip(),
        str(client_data.get("Broker", "")).strip(),
        str(client_data.get("algo", "")).strip(),
        str(client_data.get("Running Type", "")).strip(),
        "",  # Operator Name: system-assigned via server mapping, not set on add
        str(client_data.get("Category", "")).strip(),
        str(client_data.get("SubCategory", "")).strip(),
        str(client_data.get("Acc Type", "")).strip(),
    )

    with connection.cursor() as cursor:
        cursor.execute(query, values)


def update_client(
    connection: pymysql.connections.Connection,
    client_data: dict,
) -> None:
    """Updates an existing client row identified by userId."""

    query = """
        UPDATE clients
        SET
            alias          = %s,
            Broker         = %s,
            algo           = %s,
            `Running Type`   = %s,
            Category       = %s,
            SubCategory    = %s,
            `Acc Type`       = %s
        WHERE userId = %s
    """
    # Note: Operator Name is NOT updated here — it is system-managed
    # via the server mapping (Tables UI → server dropdown → auto-sync).

    values = (
        str(client_data.get("alias", "")).strip(),
        str(client_data.get("Broker", "")).strip(),
        str(client_data.get("algo", "")).strip(),
        str(client_data.get("Running Type", "")).strip(),
        str(client_data.get("Category", "")).strip(),
        str(client_data.get("SubCategory", "")).strip(),
        str(client_data.get("Acc Type", "")).strip(),
        str(client_data.get("userId", "")).strip(),
    )

    with connection.cursor() as cursor:
        cursor.execute(query, values)


def upsert_client(
    connection: pymysql.connections.Connection,
    client_data: dict,
) -> str:
    """
    Inserts or updates a client by userId.
    Returns "inserted" or "updated".
    """

    user_id = str(client_data.get("userId", "")).strip()

    if client_exists_by_user_id(connection, user_id):
        update_client(connection, client_data)
        return "updated"

    insert_client(connection, client_data)
    return "inserted"


def validate_client_data(client_data: dict) -> Optional[str]:
    """
    Validates client fields.
    Required: userId, alias, Broker — must be non-empty and within max length.
    Optional: algo, Running Type, Category, SubCategory, Acc Type — length-checked only when non-empty.
    Returns None if valid, else first error message.
    """

    # Required fields — must be present and non-empty
    for field in ["userId", "alias", "Broker"]:
        error_msg = validate_string_field(
            value=str(client_data.get(field, "")),
            field_name=field,
            max_length=CLIENT_MAX_LENGTHS[field],
        )
        if error_msg:
            return error_msg

    # Optional fields — validate max length only when a value is supplied
    for field in ["algo", "Running Type", "Category", "SubCategory", "Acc Type"]:
        value = str(client_data.get(field, "")).strip()
        if value:
            max_len = CLIENT_MAX_LENGTHS[field]
            if len(value) > max_len:
                return (
                    f"Field '{field}' exceeds maximum length of {max_len} characters."
                )

    return None


# =========================
# DB: Server Operations
# =========================

def server_exists_by_name(
    connection: pymysql.connections.Connection,
    server_name: str,
) -> bool:
    """Returns True if a server_info row with the given Server name exists."""

    query = "SELECT Server FROM server_info WHERE Server = %s LIMIT 1"

    with connection.cursor() as cursor:
        cursor.execute(query, (server_name,))
        return cursor.fetchone() is not None


def insert_server(
    connection: pymysql.connections.Connection,
    server_data: dict,
) -> None:
    """Inserts a new row into server_info (form-exposed columns only)."""

    query = """
        INSERT INTO server_info
            (Server, Username, IP, Password, `Stoxxo Id`, `Stoxxo Password`,
             Algo, Expiry, Subscriptions, Logins, Active, Avlbl)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    def _safe_int(val: object, default: int = 0) -> int:
        """Casts val to int; returns default for empty/invalid values."""
        try:
            return max(0, int(str(val).strip()))
        except (ValueError, TypeError):
            return default

    values = (
        str(server_data.get("Server", "")).strip(),
        str(server_data.get("Username", "")).strip(),
        str(server_data.get("IP", "")).strip(),
        str(server_data.get("Password", "")).strip(),
        str(server_data.get("Stoxxo Id", "")).strip(),
        str(server_data.get("Stoxxo Password", "")).strip(),
        str(server_data.get("Algo", "")).strip(),
        str(server_data.get("Expiry", "")).strip(),
        _safe_int(server_data.get("Subscriptions", 0)),
        _safe_int(server_data.get("Logins", 0)),
        _safe_int(server_data.get("Active", 0)),
        _safe_int(server_data.get("Avlbl", 0)),
    )

    with connection.cursor() as cursor:
        cursor.execute(query, values)


def validate_server_data(server_data: dict) -> Optional[str]:
    """
    Validates server form fields.
    Required: Server, Username, IP, Password, Stoxxo Id, Stoxxo Password — non-empty + max length.
    Optional VARCHAR: Algo, Expiry — length-checked only when non-empty.
    Optional INT: Subscriptions, Logins, Active, Avlbl — validated only when non-empty; defaults to 0.
    Returns None if valid, else first error message.
    """

    # Required VARCHAR fields
    for field in ["Server", "Username", "IP", "Password", "Stoxxo Id", "Stoxxo Password"]:
        error_msg = validate_string_field(
            value=str(server_data.get(field, "")),
            field_name=field,
            max_length=SERVER_VARCHAR_MAX_LENGTHS[field],
        )
        if error_msg:
            return error_msg

    # Optional VARCHAR fields — length check only when provided
    for field in ["Algo", "Expiry"]:
        value = str(server_data.get(field, "")).strip()
        if value:
            max_len = SERVER_VARCHAR_MAX_LENGTHS[field]
            if len(value) > max_len:
                return (
                    f"Field '{field}' exceeds maximum length of {max_len} characters."
                )

    # Optional INT fields — validate format only when provided; empty → defaults to 0
    for field in SERVER_INT_COLUMNS:
        raw_value = str(server_data.get(field, "")).strip()
        if raw_value:
            try:
                int_value = int(raw_value)
                if int_value < 0:
                    return f"Field '{field}' must be a non-negative integer."
            except ValueError:
                return f"Field '{field}' must be a valid integer (got: '{raw_value}')."

    return None


# =========================
# Routes: Public
# =========================

@app.route("/", methods=["GET"])
def index():
    """Root route: redirect to dashboard if logged in, else show login."""

    if session.get("is_logged_in"):
        return redirect_user_by_role(session.get("role", ""))

    log_info(module=MODULE_NAME, action="index",
             message="Login page requested from root URL", status="SUCCESS")

    return render_template("login.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    """
    GET:  Show login page (redirect if already logged in).
    POST: Authenticate email/password against role_login.
    """

    if request.method == "GET":
        if session.get("is_logged_in"):
            return redirect_user_by_role(session.get("role", ""))

        log_info(module=MODULE_NAME, action="login_get",
                 message="Login page requested", status="SUCCESS")
        return render_template("login.html")

    try:
        email    = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()

        log_info(module=MODULE_NAME, action="login_post",
                 message=f"Login attempt for email: {email}", status="STARTED")

        if not email or not password:
            log_warning(module=MODULE_NAME, action="login_validation",
                        message="Email or password missing", status="FAILED")
            return render_template("login.html", error="Email and password are required.")

        user = get_user_by_email(email)

        if user is None:
            log_warning(module=MODULE_NAME, action="login_authentication",
                        message=f"Login failed. Email not found: {email}", status="FAILED")
            return render_template("login.html", error="Invalid email or password.")

        if password != str(user.get("password", "")).strip():
            log_warning(module=MODULE_NAME, action="login_authentication",
                        message=f"Login failed. Password mismatch for: {email}", status="FAILED")
            return render_template("login.html", error="Invalid email or password.")

        role = str(user.get("role", "")).strip().lower()

        if not is_valid_role(role):
            log_warning(module=MODULE_NAME, action="login_role_validation",
                        message=f"Invalid role '{role}' for: {email}", status="FAILED")
            return render_template("login.html", error="Invalid role assigned to this user.")

        session.clear()
        session["is_logged_in"] = True
        session["role"]         = role
        session["name"]         = str(user.get("name", "")).strip()
        session["ops_name"]     = str(user.get("ops_name", "")).strip()
        session["email"]        = str(user.get("email", "")).strip()

        log_info(module=MODULE_NAME, action="login_authentication",
                 message=f"Login successful for: {email}, role: {role}", status="SUCCESS")

        return redirect_user_by_role(role)

    except Exception as exc:
        log_error(module=MODULE_NAME, action="login_post",
                  message="Login request failed", error=exc, status="FAILED")
        return render_template("login.html", error="Something went wrong. Please try again.")


@app.route("/logout", methods=["GET"])
def logout():
    """Clears session and redirects to login."""

    user_email = session.get("email", "unknown")
    session.clear()

    log_info(module=MODULE_NAME, action="logout",
             message=f"User logged out: {user_email}", status="SUCCESS")

    return redirect(url_for("login"))


# =========================
# Routes: Dashboards
# =========================

@app.route("/superadmin/dashboard", methods=["GET"])
@login_required
@role_required("superadmin")
def superadmin_dashboard():
    log_info(module=MODULE_NAME, action="superadmin_dashboard",
             message="Superadmin dashboard requested", status="SUCCESS")

    return render_template(
        "superadmin/dashboard.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


@app.route("/admin/dashboard", methods=["GET"])
@login_required
@role_required("admin")
def admin_dashboard():
    log_info(module=MODULE_NAME, action="admin_dashboard",
             message="Admin dashboard requested", status="SUCCESS")

    return render_template(
        "admin/dashboard.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


@app.route("/data/dashboard", methods=["GET"])
@login_required
@role_required("data")
def data_dashboard():
    log_info(module=MODULE_NAME, action="data_dashboard",
             message="Data dashboard requested", status="SUCCESS")

    return render_template(
        "data/dashboard.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


@app.route("/operator/dashboard", methods=["GET"])
@login_required
@role_required("operator")
def operator_dashboard():
    log_info(module=MODULE_NAME, action="operator_dashboard",
             message="Operator dashboard requested", status="SUCCESS")

    return render_template(
        "operator/dashboard.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


# =========================
# Routes: Admin Panel — User
# =========================

@app.route("/admin-panel/add-user", methods=["POST"])
@login_required
def add_user():
    """
    Inserts a new user into role_login.
    Access: superadmin only.
    Form fields: role, name, ops_name, email, password
    Returns JSON.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    current_role = session.get("role", "")

    # Only superadmin can add users
    if current_role != "superadmin":
        return jsonify({"success": False, "error": "Only superadmin can add users."}), 403

    try:
        new_role = str(request.form.get("role", "")).strip().lower()
        name     = str(request.form.get("name", "")).strip()
        ops_name = str(request.form.get("ops_name", "")).strip()
        email    = str(request.form.get("email", "")).strip()
        password = str(request.form.get("password", "")).strip()

        log_info(module=MODULE_NAME, action="add_user",
                 message=f"Add user by {session.get('email')}: ops_name={ops_name}, role={new_role}",
                 status="STARTED")

        if not new_role:
            return jsonify({"success": False, "error": "Role is required."}), 400

        allowed_new_roles = get_allowed_new_roles(current_role)

        if new_role not in allowed_new_roles:
            log_warning(module=MODULE_NAME, action="add_user",
                        message=f"Unauthorized role assignment '{new_role}' by {session.get('email')}",
                        status="FAILED")
            return jsonify({
                "success": False,
                "error": f"You are not permitted to assign the role '{new_role}'.",
            }), 403

        for value, field_name, max_len in [
            (name,     "name",     ROLE_LOGIN_MAX_LENGTHS["name"]),
            (ops_name, "ops_name", ROLE_LOGIN_MAX_LENGTHS["ops_name"]),
            (email,    "email",    ROLE_LOGIN_MAX_LENGTHS["email"]),
            (password, "password", ROLE_LOGIN_MAX_LENGTHS["password"]),
        ]:
            err = validate_string_field(value, field_name, max_len)
            if err:
                return jsonify({"success": False, "error": err}), 400

        connection = get_db_connection()

        try:
            duplicate = user_exists_by_ops_name_or_email(connection, ops_name, email)

            if duplicate:
                conflict = "ops_name" if duplicate.get("ops_name") == ops_name else "email"
                return jsonify({
                    "success": False,
                    "error": f"A user with this {conflict} already exists.",
                }), 409

            insert_role_login_user(connection, {
                "role": new_role, "name": name,
                "ops_name": ops_name, "email": email, "password": password,
            })

        finally:
            connection.close()

        log_added(module=MODULE_NAME, action="add_user",
                  message=f"User '{ops_name}' added by {session.get('email')}, role: {new_role}",
                  status="SUCCESS")

        return jsonify({
            "success": True,
            "message": f"User '{name}' added successfully with role '{new_role}'.",
        }), 201

    except Exception as exc:
        log_error(module=MODULE_NAME, action="add_user",
                  message="Add user failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Something went wrong. Please try again."}), 500


# =========================
# Routes: Admin Panel — Client
# =========================

@app.route("/admin-panel/add-client", methods=["POST"])
@login_required
def add_client():
    """
    Upserts a single client by userId.
    Access: superadmin + admin.
    Returns JSON with action: inserted | updated.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        client_data: dict = {
            col: str(request.form.get(col, "")).strip()
            for col in CLIENT_FORM_COLUMNS
        }

        log_info(module=MODULE_NAME, action="add_client",
                 message=f"Add client by {session.get('email')}: userId={client_data.get('userId', '')}",
                 status="STARTED")

        validation_error = validate_client_data(client_data)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        connection = get_db_connection()

        try:
            action = upsert_client(connection, client_data)
        finally:
            connection.close()

        action_word = "added" if action == "inserted" else "updated"

        _log_fn = log_added if action == "inserted" else log_update
        _log_fn(module=MODULE_NAME, action="add_client",
                message=f"Client '{client_data['userId']}' {action_word} by {session.get('email')}",
                status="SUCCESS")

        return jsonify({
            "success": True,
            "action": action,
            "message": f"Client '{client_data['userId']}' {action_word} successfully.",
        }), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="add_client",
                  message="Add client failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Something went wrong. Please try again."}), 500


@app.route("/admin-panel/bulk-upload-clients", methods=["POST"])
@login_required
def bulk_upload_clients():
    """
    Bulk upserts clients from .xlsx. Existing userId rows are updated.
    Access: superadmin + admin.
    Returns JSON with inserted/updated/error counts.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        uploaded_file = request.files.get("file")

        if not uploaded_file or not uploaded_file.filename:
            return jsonify({"success": False, "error": "No file uploaded."}), 400

        if not uploaded_file.filename.lower().endswith(".xlsx"):
            return jsonify({"success": False, "error": "Only .xlsx files are accepted."}), 400

        log_info(module=MODULE_NAME, action="bulk_upload_clients",
                 message=f"Bulk client upload by {session.get('email')}: {uploaded_file.filename}",
                 status="STARTED")

        try:
            workbook  = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()),
                                               read_only=True, data_only=True)
        except Exception:
            return jsonify({"success": False,
                            "error": "Could not read file. Ensure it is a valid .xlsx."}), 400

        worksheet = workbook.active
        rows      = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        if not rows:
            return jsonify({"success": False, "error": "Uploaded file is empty."}), 400

        raw_headers    = [str(h).strip() if h is not None else "" for h in rows[0]]
        missing_cols   = [c for c in CLIENT_REQUIRED_BULK_COLUMNS if c not in raw_headers]

        if missing_cols:
            return jsonify({
                "success": False,
                "error": (
                    f"Missing required columns in file: {', '.join(missing_cols)}. "
                    f"Required: {', '.join(CLIENT_REQUIRED_BULK_COLUMNS)}."
                ),
            }), 400

        data_rows = rows[1:]

        if not data_rows:
            return jsonify({"success": False, "error": "No data rows found."}), 400

        inserted_count = 0
        updated_count  = 0
        error_rows: list[dict] = []

        connection = get_db_connection()

        try:
            for row_idx, row in enumerate(data_rows, start=2):
                row_values  = [str(v).strip() if v is not None else "" for v in row]
                row_dict    = dict(zip(raw_headers, row_values))
                client_data = {col: row_dict.get(col, "") for col in CLIENT_FORM_COLUMNS}

                err = validate_client_data(client_data)
                if err:
                    error_rows.append({"row": row_idx, "userId": client_data.get("userId", ""), "error": err})
                    continue

                action = upsert_client(connection, client_data)
                if action == "inserted":
                    inserted_count += 1
                else:
                    updated_count += 1

        finally:
            connection.close()

        log_info(module=MODULE_NAME, action="bulk_upload_clients",
                 message=(f"Bulk client upload by {session.get('email')}: "
                          f"inserted={inserted_count}, updated={updated_count}, errors={len(error_rows)}"),
                 status="SUCCESS")

        return jsonify({
            "success": True,
            "inserted": inserted_count,
            "updated": updated_count,
            "errors": error_rows,
            "message": (
                f"Bulk upload complete — {inserted_count} inserted, {updated_count} updated"
                + (f", {len(error_rows)} row(s) with errors." if error_rows else ".")
            ),
        }), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="bulk_upload_clients",
                  message="Bulk client upload failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Something went wrong. Please try again."}), 500


@app.route("/admin-panel/download-client-template", methods=["GET"])
@login_required
def download_client_template():
    """Generates and serves .xlsx client upload template. Access: superadmin + admin."""

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        ws.append(CLIENT_FORM_COLUMNS)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_info(module=MODULE_NAME, action="download_client_template",
                 message=f"Client template downloaded by {session.get('email')}", status="SUCCESS")

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="client_upload_template.xlsx",
        )

    except Exception as exc:
        log_error(module=MODULE_NAME, action="download_client_template",
                  message="Client template generation failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not generate template."}), 500



# =========================
# Routes: Admin Panel — Server
# =========================

@app.route("/admin-panel/add-server", methods=["POST"])
@login_required
def add_server():
    """
    Inserts a new server into server_info.
    Rejects if Server name already exists.
    Access: superadmin + admin.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        server_data = {
            col: str(request.form.get(col, "")).strip()
            for col in SERVER_FORM_COLUMNS
        }

        server_name = server_data.get("Server", "")

        log_info(module=MODULE_NAME, action="add_server",
                 message=f"Add server by {session.get('email')}: Server={server_name}",
                 status="STARTED")

        validation_error = validate_server_data(server_data)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        connection = get_db_connection()

        try:
            if server_exists_by_name(connection, server_name):
                return jsonify({
                    "success": False,
                    "error": f"Server '{server_name}' already exists. Duplicate servers cannot be added.",
                }), 409

            insert_server(connection, server_data)

        finally:
            connection.close()

        log_added(module=MODULE_NAME, action="add_server",
                  message=f"Server '{server_name}' added by {session.get('email')}",
                  status="SUCCESS")

        return jsonify({
            "success": True,
            "message": f"Server '{server_name}' added successfully.",
        }), 201

    except Exception as exc:
        log_error(module=MODULE_NAME, action="add_server",
                  message="Add server failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Something went wrong. Please try again."}), 500


@app.route("/admin-panel/bulk-upload-servers", methods=["POST"])
@login_required
def bulk_upload_servers():
    """
    Bulk inserts new servers from .xlsx.
    Rows where Server already exists are SKIPPED.
    Access: superadmin + admin.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        uploaded_file = request.files.get("file")

        if not uploaded_file or not uploaded_file.filename:
            return jsonify({"success": False, "error": "No file uploaded."}), 400

        if not uploaded_file.filename.lower().endswith(".xlsx"):
            return jsonify({"success": False, "error": "Only .xlsx files are accepted."}), 400

        log_info(module=MODULE_NAME, action="bulk_upload_servers",
                 message=f"Bulk server upload by {session.get('email')}: {uploaded_file.filename}",
                 status="STARTED")

        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        except Exception:
            return jsonify({"success": False,
                            "error": "Could not read file. Ensure it is a valid .xlsx."}), 400

        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        if not rows:
            return jsonify({"success": False, "error": "Uploaded file is empty."}), 400

        raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        missing_cols = [c for c in SERVER_REQUIRED_BULK_COLUMNS if c not in raw_headers]

        if missing_cols:
            return jsonify({
                "success": False,
                "error": (
                    f"Missing required columns in file: {', '.join(missing_cols)}. "
                    f"Required: {', '.join(SERVER_REQUIRED_BULK_COLUMNS)}."
                ),
            }), 400

        data_rows = rows[1:]

        if not data_rows:
            return jsonify({"success": False, "error": "No data rows found."}), 400

        inserted_count = 0
        skipped_rows = []
        error_rows = []

        connection = get_db_connection()

        try:
            for row_idx, row in enumerate(data_rows, start=2):
                row_values = [str(v).strip() if v is not None else "" for v in row]
                row_dict = dict(zip(raw_headers, row_values))
                server_data = {col: row_dict.get(col, "") for col in SERVER_FORM_COLUMNS}
                server_name = server_data.get("Server", "")

                err = validate_server_data(server_data)
                if err:
                    error_rows.append({"row": row_idx, "Server": server_name, "error": err})
                    continue

                if server_exists_by_name(connection, server_name):
                    skipped_rows.append({
                        "row": row_idx,
                        "Server": server_name,
                        "reason": "Server already exists.",
                    })
                    continue

                insert_server(connection, server_data)
                inserted_count += 1

        finally:
            connection.close()

        log_info(module=MODULE_NAME, action="bulk_upload_servers",
                 message=(f"Bulk server upload by {session.get('email')}: "
                          f"inserted={inserted_count}, skipped={len(skipped_rows)}, errors={len(error_rows)}"),
                 status="SUCCESS")

        parts = [f"{inserted_count} inserted"]
        if skipped_rows:
            parts.append(f"{len(skipped_rows)} skipped (already exist)")
        if error_rows:
            parts.append(f"{len(error_rows)} row(s) with errors")

        return jsonify({
            "success": True,
            "inserted": inserted_count,
            "skipped": skipped_rows,
            "errors": error_rows,
            "message": "Bulk server upload complete — " + ", ".join(parts) + ".",
        }), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="bulk_upload_servers",
                  message="Bulk server upload failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Something went wrong. Please try again."}), 500


@app.route("/admin-panel/download-server-template", methods=["GET"])
@login_required
def download_server_template():
    """Generates and serves .xlsx server upload template. Access: superadmin + admin."""

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servers"
        ws.append(SERVER_FORM_COLUMNS)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_info(module=MODULE_NAME, action="download_server_template",
                 message=f"Server template downloaded by {session.get('email')}", status="SUCCESS")

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="server_upload_template.xlsx",
        )

    except Exception as exc:
        log_error(module=MODULE_NAME, action="download_server_template",
                  message="Server template generation failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not generate template."}), 500




# =========================
# Constants: Server Info Table UI
# =========================

# All server_info columns in display order.
# "Server" is the primary key — shown but not editable via the table UI.
SERVER_INFO_ALL_COLUMNS: list[str] = [
    "Server", "Username", "IP", "Password",
    "Stoxxo Id", "Stoxxo Password", "Algo", "Expiry",
    "Subscriptions", "Logins", "Active", "Avlbl",
    "Dte", "Aum", "Remarks", "Operator", "Stoxxo URL",
]

# Columns that cannot be edited (PK)
SERVER_INFO_READONLY_COLS: frozenset[str] = frozenset({"Server"})

# VARCHAR editable columns with max lengths
SERVER_INFO_VARCHAR_MAX: dict[str, int] = {
    "Username":       20,
    "IP":             20,
    "Password":       30,
    "Stoxxo Id":      30,
    "Stoxxo Password": 30,
    "Algo":           10,
    "Expiry":         20,
    "Dte":            20,
    "Aum":            20,
    "Remarks":        200,
    "Operator":       30,
    "Stoxxo URL":     100,
}

# INT editable columns
SERVER_INFO_INT_COLS: frozenset[str] = frozenset({"Subscriptions", "Logins", "Active", "Avlbl"})

# Union of all editable columns
SERVER_INFO_EDITABLE: frozenset[str] = (
    frozenset(SERVER_INFO_VARCHAR_MAX.keys()) | SERVER_INFO_INT_COLS
)

# Columns whose names contain spaces or special chars and need backtick quoting
_SERVER_INFO_BACKTICK_COLS: frozenset[str] = frozenset({
    "Stoxxo Id", "Stoxxo Password", "Stoxxo URL",
})


# =========================
# DB: Server Info Table UI
# =========================

def get_all_server_info() -> list[dict]:
    """Returns all rows from server_info ordered by Server code."""

    query = "SELECT * FROM server_info ORDER BY Server"

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            return cursor.fetchall()
    finally:
        connection.close()


def validate_server_info_field(field: str, value: str) -> Optional[str]:
    """
    Validates a single server_info editable field.
    Returns None if valid, or an error message string.
    """

    if field not in SERVER_INFO_EDITABLE:
        return f"Field '{field}' is not editable."

    if field in SERVER_INFO_INT_COLS:
        if not value and value != "0":
            return f"Field '{field}' is required."
        try:
            int_val = int(value)
            if int_val < 0:
                return f"Field '{field}' must be a non-negative integer."
        except ValueError:
            return f"Field '{field}' must be a valid integer."
        return None

    # VARCHAR — length check only (empty allowed for non-required fields)
    max_len = SERVER_INFO_VARCHAR_MAX.get(field, 255)
    if len(value) > max_len:
        return f"Field '{field}' exceeds maximum length of {max_len} characters."

    return None


def update_server_info_field(server_code: str, field: str, value: str) -> None:
    """
    Updates a single editable column in server_info identified by Server code.
    INT columns are cast before storage.
    """

    if field in _SERVER_INFO_BACKTICK_COLS:
        col_sql = f"`{field}`"
    elif " " in field:
        col_sql = f"`{field}`"
    else:
        col_sql = field

    if field in SERVER_INFO_INT_COLS:
        typed_value = int(value)
    else:
        typed_value = value

    query = f"UPDATE server_info SET {col_sql} = %s WHERE Server = %s"

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query, (typed_value, server_code))
    finally:
        connection.close()


# Fields in server_info that cascade to the clients table when changed.
# Maps server_info column → clients column.
_SERVER_CASCADE_MAP: dict[str, str] = {
    "Operator": "`Operator Name`",
    "Algo":     "algo",
}


def cascade_server_info_to_clients(
    server_code: str,
    field: str,
    value: str,
) -> int:
    """
    When Operator or Algo is updated on a server_info row, propagates
    the new value to every client assigned to that server.

    Returns the number of client rows updated (0 if field not in cascade map).
    """

    client_col = _SERVER_CASCADE_MAP.get(field)
    if not client_col:
        return 0

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE clients SET {client_col} = %s WHERE server = %s",
                (value, server_code),
            )
            return cursor.rowcount
    finally:
        connection.close()


# =========================
# Routes: Server Info API
# =========================

@app.route("/api/server-info", methods=["GET"])
@login_required
def api_get_server_info():
    """Returns all server_info rows as JSON. Access: superadmin + admin."""

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        rows = get_all_server_info()

        servers = [
            {k: (v if v is not None else "") for k, v in row.items()}
            for row in rows
        ]

        log_info(module=MODULE_NAME, action="api_get_server_info",
                 message=f"Returned {len(servers)} server rows to {session.get('email')}",
                 status="SUCCESS")

        return jsonify({"success": True, "servers": servers}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_get_server_info",
                  message="Failed to fetch server_info", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not load server info."}), 500


@app.route("/api/server-info/update-field", methods=["POST"])
@login_required
def api_update_server_info_field():
    """
    Updates a single editable field on a server_info row.

    Expected form fields:
        server  (str) — Server code identifying the row (PK)
        field   (str) — column name to update
        value   (str) — new value

    Access: superadmin + admin.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        server_code = str(request.form.get("server", "")).strip()
        field       = str(request.form.get("field",  "")).strip()
        value       = str(request.form.get("value",  "")).strip()

        if not server_code:
            return jsonify({"success": False, "error": "server is required."}), 400

        if field not in SERVER_INFO_EDITABLE:
            return jsonify({"success": False,
                            "error": f"Field '{field}' is not editable."}), 400

        validation_error = validate_server_info_field(field, value)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        update_server_info_field(server_code, field, value)

        # Cascade Operator / Algo changes to all clients on this server.
        cascaded = cascade_server_info_to_clients(server_code, field, value)

        log_update(module=MODULE_NAME, action="api_update_server_info_field",
                   message=(
                       f"Server info field updated by {session.get('email')}: "
                       f"server={server_code}, field={field}"
                       + (f" | cascaded to {cascaded} client(s)" if cascaded else "")
                   ),
                   status="SUCCESS")

        return jsonify({"success": True, "message": "Saved.", "cascaded": cascaded}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_update_server_info_field",
                  message="Inline server_info update failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not save change."}), 500


# =========================
# Constants: Inline Edit
# =========================

# Columns the tables UI is allowed to update on the clients table.
# "server" also triggers an Operator Name sync from server_info.
CLIENT_INLINE_EDITABLE: frozenset[str] = frozenset({
    "alias", "server", "algo", "Running Type", "Acc Type",
})

RUNNING_TYPE_VALUES: frozenset[str] = frozenset({"INT", "POS"})

ACC_TYPE_VALUES: frozenset[str] = frozenset({
    "Client",
    "PS_Personal",
    "RD_Personal",
    "VT_Personal",
    "GB_Personal",
})

# Display order for the client table columns in the Tables UI.
CLIENT_TABLE_COLUMNS: list[str] = [
    "userId",
    "alias",
    "Broker",
    "server",
    "algo",
    "Running Type",
    "Operator Name",
    "Category",
    "SubCategory",
    "Acc Type",
]


# =========================
# DB: Tables UI Operations
# =========================

def get_all_clients() -> list[dict]:
    """
    Returns all clients joined with server_info so that Operator Name
    is always live from server_info.Operator (never a stale stored value).
    """

    query = """
        SELECT
            c.userId,
            c.alias,
            c.Broker,
            c.server,
            c.algo,
            c.`Running Type`,
            COALESCE(s.Operator, '') AS `Operator Name`,
            c.Category,
            c.SubCategory,
            c.`Acc Type`
        FROM clients c
        LEFT JOIN server_info s ON c.server = s.Server
        ORDER BY c.userId
    """

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            return cursor.fetchall()
    finally:
        connection.close()


def get_servers_for_dropdown() -> list[dict]:
    """
    Returns Server code + Operator name from server_info.
    Used to populate the server dropdown and auto-fill Operator Name.
    """

    query = "SELECT Server, Operator FROM server_info ORDER BY Server"

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            return cursor.fetchall()
    finally:
        connection.close()


def validate_inline_field(field: str, value: str) -> Optional[str]:
    """
    Validates a single inline-editable client field.
    Returns None if valid, or an error message string.
    """

    if field not in CLIENT_INLINE_EDITABLE:
        return f"Field '{field}' is not editable."

    if field == "alias":
        return validate_string_field(value, "alias", 30)

    if field == "algo":
        if len(value) > 10:
            return "Field 'algo' exceeds maximum length of 10 characters."
        return None

    if field == "Running Type":
        if value not in RUNNING_TYPE_VALUES:
            return f"Running Type must be one of: {', '.join(sorted(RUNNING_TYPE_VALUES))}."
        return None

    if field == "Acc Type":
        if value not in ACC_TYPE_VALUES:
            return f"Acc Type must be one of: {', '.join(sorted(ACC_TYPE_VALUES))}."
        return None

    if field == "server":
        # Empty string is acceptable (client has no server assigned)
        return None

    return None


def update_client_inline_field(
    user_id: str,
    field: str,
    value: str,
) -> dict:
    """
    Updates a single editable field on the clients table.

    If field == "server", also syncs Operator Name from server_info.

    Returns a dict with keys:
        operator_name (str | None) — new operator name when field=="server",
                                     else None.
    """

    connection = get_db_connection()

    try:
        if field == "server":
            # Resolve operator from server_info for the chosen server.
            operator_name: str = ""

            if value:
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT Operator FROM server_info WHERE Server = %s LIMIT 1",
                        (value,),
                    )
                    row = cursor.fetchone()
                    operator_name = str(row["Operator"]).strip() if row and row.get("Operator") else ""

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE clients
                    SET server = %s, `Operator Name` = %s
                    WHERE userId = %s
                    """,
                    (value, operator_name, user_id),
                )

            return {"operator_name": operator_name}

        else:
            # Generic column update — build query dynamically using a
            # whitelist-derived column name (safe: field already validated).
            col = f"`{field}`" if " " in field else field

            with connection.cursor() as cursor:
                cursor.execute(
                    f"UPDATE clients SET {col} = %s WHERE userId = %s",
                    (value, user_id),
                )

            return {"operator_name": None}

    finally:
        connection.close()


# =========================
# Routes: Admin Control Pages
# =========================

@app.route("/superadmin/admin-control", methods=["GET"])
@login_required
@role_required("superadmin")
def superadmin_admin_control():
    log_info(module=MODULE_NAME, action="superadmin_admin_control",
             message="Superadmin admin control page requested", status="SUCCESS")

    return render_template(
        "superadmin/admin_control.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


@app.route("/admin/admin-control", methods=["GET"])
@login_required
@role_required("admin")
def admin_admin_control():
    log_info(module=MODULE_NAME, action="admin_admin_control",
             message="Admin admin control page requested", status="SUCCESS")

    return render_template(
        "admin/admin_control.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


# =========================
# Routes: Tables Pages
# =========================

@app.route("/superadmin/tables", methods=["GET"])
@login_required
@role_required("superadmin")
def superadmin_tables():
    log_info(module=MODULE_NAME, action="superadmin_tables",
             message="Superadmin tables page requested", status="SUCCESS")

    return render_template(
        "superadmin/tables.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


@app.route("/admin/tables", methods=["GET"])
@login_required
@role_required("admin")
def admin_tables():
    log_info(module=MODULE_NAME, action="admin_tables",
             message="Admin tables page requested", status="SUCCESS")

    return render_template(
        "admin/tables.html",
        name=session.get("name"),
        ops_name=session.get("ops_name"),
        email=session.get("email"),
        role=session.get("role"),
    )


# =========================
# Routes: Tables API
# =========================

@app.route("/api/clients", methods=["GET"])
@login_required
def api_get_clients():
    """Returns all clients as JSON. Access: superadmin + admin."""

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        rows = get_all_clients()

        # Normalise: replace None with empty string for JSON cleanliness
        clients = [
            {k: (v if v is not None else "") for k, v in row.items()}
            for row in rows
        ]

        log_info(module=MODULE_NAME, action="api_get_clients",
                 message=f"Returned {len(clients)} clients to {session.get('email')}",
                 status="SUCCESS")

        return jsonify({"success": True, "clients": clients}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_get_clients",
                  message="Failed to fetch clients", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not load clients."}), 500


@app.route("/api/operators", methods=["GET"])
@login_required
def api_get_operators():
    """
    Returns ops_name list of all users with role='operator' from role_login.
    Used to populate the Operator dropdown in server_info table.
    Access: superadmin + admin.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT ops_name FROM role_login WHERE role = 'operator' ORDER BY ops_name"
                )
                rows = cursor.fetchall()
        finally:
            connection.close()

        operators = [str(r["ops_name"]).strip() for r in rows if r.get("ops_name")]

        return jsonify({"success": True, "operators": operators}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_get_operators",
                  message="Failed to fetch operators", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not load operators."}), 500


@app.route("/api/servers", methods=["GET"])
@login_required
def api_get_servers():
    """Returns server codes + operators from server_info. Access: superadmin + admin."""

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        rows = get_servers_for_dropdown()
        servers = [
            {k: (v if v is not None else "") for k, v in row.items()}
            for row in rows
        ]

        log_info(module=MODULE_NAME, action="api_get_servers",
                 message=f"Returned {len(servers)} server records to {session.get('email')}",
                 status="SUCCESS")

        return jsonify({"success": True, "servers": servers}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_get_servers",
                  message="Failed to fetch servers", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not load servers."}), 500


@app.route("/api/clients/update-field", methods=["POST"])
@login_required
def api_update_client_field():
    """
    Updates a single inline-editable field on a client record.

    Expected form fields:
        userId  (str) — identifies the row
        field   (str) — column name to update
        value   (str) — new value

    When field == "server", also syncs Operator Name from server_info.

    Returns JSON with success flag + optional operator_name.
    Access: superadmin + admin.
    """

    access_error = check_admin_panel_access()
    if access_error:
        return access_error

    try:
        user_id = str(request.form.get("userId", "")).strip()
        field   = str(request.form.get("field",  "")).strip()
        value   = str(request.form.get("value",  "")).strip()

        if not user_id:
            return jsonify({"success": False, "error": "userId is required."}), 400

        if field not in CLIENT_INLINE_EDITABLE:
            return jsonify({"success": False, "error": f"Field '{field}' is not editable."}), 400

        validation_error = validate_inline_field(field, value)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        result = update_client_inline_field(user_id, field, value)

        log_update(module=MODULE_NAME, action="api_update_client_field",
                   message=(
                       f"Client field updated by {session.get('email')}: "
                       f"userId={user_id}, field={field}"
                   ),
                   status="SUCCESS")

        return jsonify({
            "success": True,
            "message": "Saved.",
            "operator_name": result.get("operator_name"),
        }), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_update_client_field",
                  message="Inline client update failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not save change."}), 500


# =========================
# Constants: Users Management (superadmin only)
# =========================

# Columns shown in the users management table
USER_TABLE_COLUMNS: list[str] = ["email", "role", "name", "ops_name", "password"]

# Editable columns + their max lengths (email is the PK — shown read-only)
USER_EDITABLE_FIELDS: dict[str, int] = {
    "role":     20,
    "name":     30,
    "ops_name": 20,
    "password": 30,
}

VALID_ASSIGNABLE_ROLES: frozenset[str] = frozenset(
    {"superadmin", "admin", "data", "operator"}
)


# =========================
# DB: Users Management
# =========================

def get_all_users() -> list[dict]:
    """Returns all rows from role_login ordered by role then name."""

    query = """
        SELECT email, role, name, ops_name, password
        FROM role_login
        ORDER BY role, name
    """

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            return cursor.fetchall()
    finally:
        connection.close()


def update_user_field(email: str, field: str, value: str) -> None:
    """Updates a single editable column in role_login identified by email."""

    # field is already validated against USER_EDITABLE_FIELDS whitelist before call
    query = f"UPDATE role_login SET `{field}` = %s WHERE email = %s"

    connection = get_db_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(query, (value, email))
    finally:
        connection.close()


# =========================
# Routes: Users Management API (superadmin only)
# =========================

@app.route("/api/users", methods=["GET"])
@login_required
def api_get_users():
    """Returns all role_login rows as JSON. Superadmin only."""

    if session.get("role") != "superadmin":
        return jsonify({"success": False, "error": "Permission denied."}), 403

    try:
        rows = get_all_users()
        users = [
            {k: (v if v is not None else "") for k, v in row.items()}
            for row in rows
        ]

        log_info(module=MODULE_NAME, action="api_get_users",
                 message=f"Returned {len(users)} user rows to {session.get('email')}",
                 status="SUCCESS")

        return jsonify({"success": True, "users": users}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_get_users",
                  message="Failed to fetch users", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not load users."}), 500


@app.route("/api/users/update-field", methods=["POST"])
@login_required
def api_update_user_field():
    """
    Updates a single editable field on a role_login row.

    Form fields:
        email  (str) — identifies the row (PK)
        field  (str) — column name to update
        value  (str) — new value

    Superadmin only.
    """

    if session.get("role") != "superadmin":
        return jsonify({"success": False, "error": "Permission denied."}), 403

    try:
        email = str(request.form.get("email", "")).strip()
        field = str(request.form.get("field", "")).strip()
        value = str(request.form.get("value", "")).strip()

        if not email:
            return jsonify({"success": False, "error": "email is required."}), 400

        if field not in USER_EDITABLE_FIELDS:
            return jsonify({"success": False,
                            "error": f"Field '{field}' is not editable."}), 400

        if field == "role":
            if value not in VALID_ASSIGNABLE_ROLES:
                return jsonify({
                    "success": False,
                    "error": f"Invalid role '{value}'.",
                }), 400
        else:
            max_len = USER_EDITABLE_FIELDS[field]
            if not value:
                return jsonify({"success": False,
                                "error": f"Field '{field}' cannot be empty."}), 400
            if len(value) > max_len:
                return jsonify({
                    "success": False,
                    "error": f"Field '{field}' exceeds max length of {max_len}.",
                }), 400

        update_user_field(email, field, value)

        log_update(module=MODULE_NAME, action="api_update_user_field",
                   message=(
                       f"User field updated by {session.get('email')}: "
                       f"email={email}, field={field}"
                   ),
                   status="SUCCESS")

        return jsonify({"success": True, "message": "Saved."}), 200

    except Exception as exc:
        log_error(module=MODULE_NAME, action="api_update_user_field",
                  message="Inline user update failed", error=exc, status="FAILED")
        return jsonify({"success": False, "error": "Could not save change."}), 500


# =========================
# Error Handlers
# =========================


@app.errorhandler(404)
def page_not_found(error):
    log_error(module=MODULE_NAME, action="404_error",
              message="Page not found", error=error, status="FAILED")
    return render_template("login.html", error="Page not found."), 404


@app.errorhandler(500)
def internal_server_error(error):
    log_error(module=MODULE_NAME, action="500_error",
              message="Internal server error", error=error, status="FAILED")
    return render_template("login.html", error="Internal server error."), 500
